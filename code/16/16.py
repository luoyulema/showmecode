import chardet
from openpyxl import Workbook
import json
import os


def turntoxls(path):
    file = open(path, 'r+').read()
    filecode = chardet.detect(file)
    code = filecode['encoding']
    filejson = json.loads(file, encoding=code)
    print filejson
    wb = Workbook()
    ws = wb.active

    if os.path.exists('city.xls'):
        os.remove('city.xls')
    for i in range(1, len(filejson)):
        ws.cell(row=i, column=1).value = i
        ws.cell(row=i, column=2).value = filejson[str(i)]
    wb.save('city.xls')

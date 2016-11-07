from openpyxl import Workbook
import json
import os


def turntoxls(path):
    file = open(path, 'r').read()
    filetext = json.loads(file, encoding='gb2312')
    filesort = sorted(filetext.items(), key=lambda t: t[0])
    wb = Workbook()
    ws = wb.active
    for i in range(0, len(filesort)):
        ws.cell(row=i+1, column=1).value = i + 1
        for k in range(0, len(filesort[i])):
            for v in range(0, len(filesort[i][k])):
                ws.cell(row=i + 1, column=v + 2).value = filesort[i][k][v]
    if os.path.exists('student.xls'):
        os.remove('student.xls')
    wb.save('student.xls')
